import openpyxl
from openpyxl.styles import Border, Side, PatternFill
import webcolors
from colorama import Fore, Style
import os
import csv
import time
from tqdm import tqdm


# Function to create a new Excel workbook
def create_workbook():
    workbook = openpyxl.Workbook()
    return workbook


# Function to add a new sheet to the workbook
def add_sheet(workbook):
    sheet = workbook.active
    return sheet


# Function to get column values from the user
def get_column_values(sheet, column_index):
    values = []
    while True:
        value = input(Fore.MAGENTA + Style.BRIGHT + f"Enter a value for column {chr(ord('A')+column_index)} (press Enter to finish): ")
        if value == "":
            print()
            break
        values.append(value)
    return values


# Function to write column values to the worksheet
def write_column_values(sheet, column_index, values):
    for row_index, value in enumerate(values, start=1):
        sheet.cell(row=row_index, column=column_index + 1, value=value)

# Function to set column width based on content
def set_column_width(sheet, column_index, max_length):
    column_letter = chr(ord('A') + column_index)
    sheet.column_dimensions[column_letter].width = max_length + 2  # Add some extra padding


# Function to get RGB color from the user
def get_color_rgb(color_name):
    while True:
        try:
            color_rgb = webcolors.name_to_rgb(color_name)
            return color_rgb
        except ValueError:
            print(Fore.RED + Style.BRIGHT + "Invalid color name entered.")
            color_name = input(Fore.LIGHTCYAN_EX + Style.BRIGHT + "Please enter a valid color name: ")


# Function to convert RGB to hexadecimal color
def convert_rgb_to_hex(rgb):
    return f"{rgb.red:02X}{rgb.green:02X}{rgb.blue:02X}"


# Function to apply fill color to a range of cells
def apply_fill_color(sheet, fill_color_hex, min_row, max_row, min_col, max_col):
    fill = PatternFill(start_color=fill_color_hex, end_color=fill_color_hex, fill_type="solid")
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.fill = fill


# Function to apply borders to a range of cells
def apply_borders(sheet, min_row, max_row, min_col, max_col):
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                          bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_style


# Function to save the workbook to a file
def save_workbook(workbook, file_name):
    workbook.save(file_name)


# Function to open the saved file
def open_file(file_name):
    os.startfile(file_name)


# Function to generate the Excel document
def generate_excel():
    workbook = create_workbook()
    sheet = add_sheet(workbook)

    while True:
        try:
            choice = input(Fore.YELLOW + Style.BRIGHT + "Do you want to enter data manually (M) or import from a CSV file (C)? ").strip().upper()
            if choice == "M":
                clm = int(input(Fore.YELLOW + Style.BRIGHT + 'Enter the number of columns to be used: '))
                break
            elif choice == "C":
                csv_file_path = input(Fore.YELLOW + Style.BRIGHT + "Enter the path of the CSV file: ")
                with open(csv_file_path, 'r') as csv_file:
                    csv_reader = csv.reader(csv_file)
                    first_row = next(csv_reader)
                    clm = len(first_row)
                break
            else:
                print(Fore.RED + "Invalid choice. Please enter 'M' or 'C'.")
        except ValueError:
            print(Fore.RED + "Invalid input. Please enter a valid choice.")

    print()
    if choice == "M":
        for i in range(clm):
            values = get_column_values(sheet, i)
            write_column_values(sheet, i, values)

            max_length = max(len(str(value)) for value in values)
            set_column_width(sheet, i, max_length)
    elif choice == "C":
        with open(csv_file_path, 'r') as csv_file:
            csv_reader = csv.reader(csv_file)
            for i, row in enumerate(csv_reader):
                write_column_values(sheet, i, row)
                max_length = max(len(str(value)) for value in row)
                set_column_width(sheet, i, max_length)

    while True:
        try:
            first_row_color = input(Fore.LIGHTCYAN_EX + Style.BRIGHT + "Enter the color for the first row: ")
            first_row_color_rgb = get_color_rgb(first_row_color)
            break
        except ValueError:
            print(Fore.RED + Style.BRIGHT + "Invalid color name entered.")

    while True:
        try:
            remaining_cells_color = input(Fore.LIGHTCYAN_EX + Style.BRIGHT + "Enter the color for the remaining cells: ")
            remaining_cells_color_rgb = get_color_rgb(remaining_cells_color)
            break
        except ValueError:
            print(Fore.RED + Style.BRIGHT + "Invalid color name entered.")

    first_row_color_hex = convert_rgb_to_hex(first_row_color_rgb)
    remaining_cells_color_hex = convert_rgb_to_hex(remaining_cells_color_rgb)

    apply_fill_color(sheet, first_row_color_hex, 1, 1, 1, sheet.max_column)
    apply_fill_color(sheet, remaining_cells_color_hex, 2, sheet.max_row, 1, sheet.max_column)

    apply_borders(sheet, 1, sheet.max_row, 1, sheet.max_column)

    file_name = f'' #file path in this format (keep the variable) C:\\Users\\OneDrive\\Desktop\\code\\{Name} - Annual Performance.xlsx
    print()
    workbook.save(file_name)
    print()
    print(Fore.RED + Style.BRIGHT + f'''Your excel sheet - {Name}-Annual Performance
has been made and saved in the \code folder''')
    print()
    time.sleep(2)
    print(Fore.RED + Style.BRIGHT + "Opening Document....")
    os.startfile(file_name)

print(Fore.LIGHTBLUE_EX + Style.BRIGHT + '''Hey want to create an Excel document of your annual sales report
Dont worry we got you!!!
Just type the required fields below.''')
print()
Name = input(Fore.LIGHTCYAN_EX + Style.BRIGHT + 'Enter the name of the company - ')
print()
print(Fore.LIGHTYELLOW_EX + Style.BRIGHT + 'The Table')
generate_excel()
